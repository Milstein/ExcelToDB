from openpyxl import load_workbook

wb = load_workbook('formula.xlsx', data_only=True)

ws = wb.active

#print(ws['C2'].value)
print(ws.cell(row=2, column=1).value)




